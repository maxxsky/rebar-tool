"""Export Excel — 4 sheet (F3). Spec 04-SPEC-output.md.

Traceability di setiap sheet: nilai config aktual yang dipakai.
Panjang tampil meter 3 desimal; internal tetap mm integer.
File baru tiap run (timestamp) — tidak menimpa.
"""

import hashlib
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models import OptimizeResult, ProjectConfig

TOOL_VERSION = "0.1.0"

# warna
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")   # kuning warning
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
SUB_FILL = PatternFill("solid", fgColor="D6E4F0")
GRAND_FILL = PatternFill("solid", fgColor="FFE699")
POTONG_FILL = PatternFill("solid", fgColor="DDEBF7")  # biru muda potongan
SISA_SIMPAN_FILL = PatternFill("solid", fgColor="D9D9D9")  # abu reusable
SISA_BUANG_FILL = PatternFill("solid", fgColor="FCE4EC")   # merah muda buang
BORDER_SEL = 250   # 1 sel ≈ 250 mm di bar pola


def _hash_file(path: Path) -> str:
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()[:16]


def _fmt_m(v: int) -> float:
    """mm → meter, 3 desimal."""
    return round(v / 1000.0, 3)


def _header(ws, cfg, warnings=(), extra_lines=()):
    """Header traceability baris 1-8 — nilai config aktual."""
    ws["A1"] = f"PROYEK  : {cfg.nama} ({cfg.kode})"
    ws["A2"] = f"SUMBER  : {cfg.sumber.dokumen} {cfg.sumber.revisi} ({cfg.sumber.tanggal})"
    ws["A3"] = f"CATATAN : {cfg.sumber.catatan}"
    ws["A4"] = (f"DIBUAT  : {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
                f"rebar-tool v{TOOL_VERSION}")
    ws["A5"] = ("─" * 72)
    cover_str = " ".join(f"{k} {v}" for k, v in sorted(cfg.cover.items()))
    ld_str = " ".join(f"D{k}={v}" for k, v in sorted(cfg.ld.items()))
    hook_str = " ".join(
        f"{sudut}°:D{k}={v}" for sudut, m in sorted(cfg.hook_tail.items())
        for k, v in sorted(m.items()))
    ws["A6"] = (f"PARAMETER: cover {cover_str} | kerf {cfg.stok.kerf_mm}mm | "
                f"batang {cfg.stok.panjang_batang_mm}mm")
    ws["A7"] = f"           Ld: {ld_str}"
    ws["A8"] = f"           hook tail: {hook_str}"
    r = 9
    for line in extra_lines:
        ws[f"A{r}"] = line
        r += 1
    for w in warnings:
        cell = ws[f"A{r}"]
        cell.value = f"⚠ WARNING: {w}"
        cell.fill = WARN_FILL
        r += 1
    ws[f"A{r}"] = "─" * 72
    return r + 1


def _agregat_berat_per_dia(cuts, cfg):
    """dict dia -> {panjang_m, jumlah, berat_kg}."""
    out = {}
    for c in cuts:
        d = out.setdefault(c.dia, {"panjang_total_m": 0.0, "berat_kg": 0.0,
                                   "jumlah": 0})
        d["panjang_total_m"] += _fmt_m(c.panjang_mm) * c.jumlah
        d["berat_kg"] += _fmt_m(c.panjang_mm) * c.jumlah * cfg.unit_weight[c.dia]
        d["jumlah"] += c.jumlah
    return out


def generate_excel(cfg: ProjectConfig, elemen_list, cuts_bbs,
                   hasil_opt: dict[int, OptimizeResult], config_dir: Path,
                   out_path: Path) -> Path:
    """cuts_bbs = daftar Cut dari generate_bbs (per bar mark, sebelum agregasi)."""
    warnings = list(cfg.warnings)
    wb = Workbook()

    # ── Sheet 1: BBS ────────────────────────────────────────
    ws = wb.active
    ws.title = "BBS"
    r = _header(ws, cfg, warnings)
    headers = ["Bar Mark", "Lokasi", "Tipe", "Posisi", "Shape", "Dia (mm)",
               "Panjang (m)", "Jml/Elem", "Jml Elemen", "Total Batang",
               "Total Panjang (m)", "Unit Wt (kg/m)", "Total Berat (kg)"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(r, i, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    r += 1

    # urut per tipe elemen, lalu per posisi — cara orang baca gambar
    by_dia = {}
    cuts_sorted = sorted(cuts_bbs, key=lambda c: (c.tipe_elemen, c.posisi,
                                                  c.dia, -c.panjang_mm))
    for c in cuts_sorted:
        by_dia.setdefault(c.dia, []).append(c)
        panjang_m = _fmt_m(c.panjang_mm)
        ws.cell(r, 1, c.bar_mark)
        ws.cell(r, 2, c.lokasi)
        ws.cell(r, 3, c.tipe_elemen)
        ws.cell(r, 4, c.posisi)
        ws.cell(r, 5, c.shape_code)
        ws.cell(r, 6, c.dia)
        ws.cell(r, 7, panjang_m)
        ws.cell(r, 8, "—")                       # jml/elem — tidak tersimpan di Cut agregat
        ws.cell(r, 9, "—")                       # jml elemen — lihat bar_mark & lokasi
        ws.cell(r, 10, c.jumlah)
        ws.cell(r, 11, round(panjang_m * c.jumlah, 3))
        ws.cell(r, 12, cfg.unit_weight[c.dia])
        ws.cell(r, 13, round(panjang_m * c.jumlah * cfg.unit_weight[c.dia], 2))
        r += 1

    # subtotal per diameter
    for dia in sorted(by_dia):
        sub = _agregat_berat_per_dia(by_dia[dia], cfg)[dia]
        ws.cell(r, 1, f"SUBTOTAL D{dia}")
        ws.cell(r, 10, sub["jumlah"])
        ws.cell(r, 11, round(sub["panjang_total_m"], 3))
        ws.cell(r, 13, round(sub["berat_kg"], 2))
        for col in range(1, 14):
            ws.cell(r, col).fill = SUB_FILL
        r += 1

    # grand total
    total_jumlah = sum(c.jumlah for c in cuts_bbs)
    total_panjang = sum(_fmt_m(c.panjang_mm) * c.jumlah for c in cuts_bbs)
    total_berat = sum(_fmt_m(c.panjang_mm) * c.jumlah * cfg.unit_weight[c.dia]
                      for c in cuts_bbs)
    ws.cell(r, 1, "GRAND TOTAL")
    ws.cell(r, 10, total_jumlah)
    ws.cell(r, 11, round(total_panjang, 3))
    ws.cell(r, 13, round(total_berat, 2))
    for col in range(1, 14):
        ws.cell(r, col).fill = GRAND_FILL
    for col, w in zip(range(1, 14), [14, 20, 8, 9, 8, 8, 12, 9, 11, 12,
                                     16, 13, 15]):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 2: POLA POTONG ────────────────────────────────
    ws2 = wb.create_sheet("POLA POTONG")
    r = _header(ws2, cfg, warnings,
                extra_lines=("CATATAN: pola diurutkan frekuensi tertinggi. "
                             "Jumlah pola belum dioptimasi utk kemudahan "
                             "lapangan (PATCH-01).",))
    for dia in sorted(hasil_opt):
        res = hasil_opt[dia]
        ws2.cell(r, 1, f"DIAMETER D{dia}   |   Batang stok {cfg.stok.panjang_batang_mm} mm")
        ws2.cell(r, 1).font = Font(bold=True, size=12)
        r += 1
        for idx, p in enumerate(res.patterns, 1):
            ws2.cell(r, 1, f"POLA {'ABCDEFGHIJKLMNOPQRSTUVWXYZ'[idx - 1]}   × {p.frekuensi} batang")
            r += 1
            # bar visual: 1 sel ≈ 250 mm
            total_sel = max(1, cfg.stok.panjang_batang_mm // BORDER_SEL + 1)
            col = 1
            for pot in p.potongan:
                n_sel = max(1, round(pot / BORDER_SEL))
                ws2.merge_cells(start_row=r, start_column=col,
                                end_row=r, end_column=col + n_sel - 1)
                cell = ws2.cell(r, col, _fmt_m(pot))
                cell.fill = POTONG_FILL
                cell.alignment = Alignment(horizontal="center")
                col += n_sel
                # kerf antar potongan: 1 sel kosong tipis
                ws2.merge_cells(start_row=r, start_column=col,
                                end_row=r, end_column=col)
                ws2.cell(r, col, "")
                col += 1
            # sisa
            if col <= total_sel:
                n_sel = total_sel - col + 1
                ws2.merge_cells(start_row=r, start_column=col,
                                end_row=r, end_column=total_sel)
                sisa_label = f"sisa {p.sisa_mm} ({'simpan' if p.reusable else 'buang'})"
                cell = ws2.cell(r, col, sisa_label)
                cell.fill = SISA_SIMPAN_FILL if p.reusable else SISA_BUANG_FILL
                cell.alignment = Alignment(horizontal="center")
            r += 1
        ws2.cell(r, 1, (f"Total batang D{dia}: {res.total_batang} | "
                        f"Berat: {round(res.total_panjang_terpakai_mm / 1000 * cfg.unit_weight[dia], 1)} kg | "
                        f"Sisa buang: {round((res.total_sisa_mm - res.sisa_reusable_mm) / 1000, 2)} m "
                        f"({res.waste_pct:.2f}%) | Sisa simpan: "
                        f"{round(res.sisa_reusable_mm / 1000, 2)} m"))
        ws2.cell(r, 1).font = Font(bold=True)
        r += 2

    # ── Sheet 3: RINGKASAN ─────────────────────────────────
    ws3 = wb.create_sheet("RINGKASAN")
    r = _header(ws3, cfg, warnings)
    ws3.cell(r, 1, "A. Kebutuhan Material").font = Font(bold=True, size=11)
    r += 1
    for i, h in enumerate(["Diameter", "Jml Batang", "Panjang Total (m)",
                           "Berat (kg)", "Berat (ton)"], 1):
        ws3.cell(r, i, h).fill = HDR_FILL
        ws3.cell(r, i).font = HDR_FONT
    r += 1
    total_batang = 0
    total_panjang = 0.0
    total_berat = 0.0
    for dia in sorted(hasil_opt):
        res = hasil_opt[dia]
        panjang_m = res.total_panjang_terpakai_mm / 1000
        berat = panjang_m * cfg.unit_weight[dia]
        ws3.cell(r, 1, f"D{dia}")
        ws3.cell(r, 2, res.total_batang)
        ws3.cell(r, 3, round(panjang_m, 3))
        ws3.cell(r, 4, round(berat, 2))
        ws3.cell(r, 5, round(berat / 1000, 2))
        total_batang += res.total_batang
        total_panjang += panjang_m
        total_berat += berat
        r += 1
    ws3.cell(r, 1, "TOTAL")
    ws3.cell(r, 2, total_batang)
    ws3.cell(r, 3, round(total_panjang, 3))
    ws3.cell(r, 4, round(total_berat, 2))
    ws3.cell(r, 5, round(total_berat / 1000, 2))
    for col in range(1, 6):
        ws3.cell(r, col).fill = GRAND_FILL
    r += 2

    ws3.cell(r, 1, "B. Analisis Waste").font = Font(bold=True, size=11)
    r += 1
    for i, h in enumerate(["Diameter", "Terpakai (m)", "Kerf (m)",
                           "Sisa Buang (m)", "Sisa Simpan (m)",
                           "Waste Bersih %", "Waste Kotor %"], 1):
        ws3.cell(r, i, h).fill = HDR_FILL
        ws3.cell(r, i).font = HDR_FONT
    r += 1
    for dia in sorted(hasil_opt):
        res = hasil_opt[dia]
        ws3.cell(r, 1, f"D{dia}")
        ws3.cell(r, 2, round(res.total_panjang_terpakai_mm / 1000, 3))
        ws3.cell(r, 3, round(res.total_kerf_mm / 1000, 3))
        ws3.cell(r, 4, round((res.total_sisa_mm - res.sisa_reusable_mm) / 1000, 3))
        ws3.cell(r, 5, round(res.sisa_reusable_mm / 1000, 3))
        ws3.cell(r, 6, res.waste_pct)
        ws3.cell(r, 7, res.waste_kotor_pct)
        r += 1
    r += 1

    ws3.cell(r, 1, "C. Dampak Pembatasan Pola").font = Font(bold=True, size=11)
    r += 1
    for i, h in enumerate(["Diameter", "Pola sebelum", "Pola sesudah",
                           "Waste tanpa batas %", "Waste dengan batas %",
                           "Selisih batang"], 1):
        ws3.cell(r, i, h).fill = HDR_FILL
        ws3.cell(r, i).font = HDR_FONT
    r += 1
    for dia in sorted(hasil_opt):
        res = hasil_opt[dia]
        ws3.cell(r, 1, f"D{dia}")
        ws3.cell(r, 2, res.pola_sebelum_batasi)
        ws3.cell(r, 3, res.pola_sesudah_batasi)
        ws3.cell(r, 4, res.waste_pct_tanpa_batasi)
        ws3.cell(r, 5, res.waste_pct)
        ws3.cell(r, 6, "")
        r += 1

    # ── Sheet 4: LOG ───────────────────────────────────────
    ws4 = wb.create_sheet("LOG")
    r = 1
    ws4.cell(r, 1, "REBAR-TOOL LOG — audit").font = Font(bold=True, size=12)
    r += 2
    config_files = sorted(Path(config_dir).glob("*.yaml"))
    hashes = {p.name: _hash_file(p) for p in config_files}
    ws4.cell(r, 1, f"VERSI TOOL   : rebar-tool v{TOOL_VERSION}")
    r += 1
    ws4.cell(r, 1, f"WAKTU        : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    r += 1
    for name, h in hashes.items():
        ws4.cell(r, 1, f"HASH CONFIG  : {name} sha256={h}")
        r += 1
    ws4.cell(r, 1, f"INPUT ELEMEN : {len(elemen_list)} grup baris")
    r += 2
    ws4.cell(r, 1, "── DUMP CONFIG ──").font = Font(bold=True)
    r += 1
    for name, h in hashes.items():
        ws4.cell(r, 1, f"### {name}")
        r += 1
        for line in Path(config_dir, name).read_text().splitlines():
            ws4.cell(r, 1, line)
            r += 1
    r += 1
    ws4.cell(r, 1, "── WARNING VALIDASI ──").font = Font(bold=True)
    r += 1
    for w in warnings:
        ws4.cell(r, 1, f"⚠ {w}")
        r += 1

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
