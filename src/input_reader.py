"""Input elemen — baca input/elemen.xlsx (F2).

Kolom: tipe | bentang_bersih_mm | jumlah | lokasi
- bentang_bersih_mm = jarak bersih antar muka tumpuan, BUKAN as-ke-as.
- Fail loud: tipe tidak dikenal, bentang negatif, dst.
"""

from pathlib import Path

from openpyxl import load_workbook

from models import ConfigError, ElemenInput


def baca_elemen_xlsx(path, templates) -> list[ElemenInput]:
    path = Path(path)
    if not path.exists():
        raise ConfigError(f"File input tidak ditemukan: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # cari header
    header = [str(c.value).strip().lower() if c.value is not None else ""
              for c in ws[1]]
    try:
        i_tipe = header.index("tipe")
        i_bentang = header.index("bentang_bersih_mm")
        i_jumlah = header.index("jumlah")
        i_lokasi = header.index("lokasi") if "lokasi" in header else None
    except ValueError as e:
        raise ConfigError(
            f"Kolom wajib di elemen.xlsx: 'tipe', 'bentang_bersih_mm', 'jumlah'. "
            f"Ditemukan: {header}. ({e})")

    errors: list[str] = []
    elemen: list[ElemenInput] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[i_tipe] is None and row[i_bentang] is None:
            continue  # baris kosong
        tipe = str(row[i_tipe]).strip() if row[i_tipe] is not None else ""
        bentang_raw = row[i_bentang]
        jumlah_raw = row[i_jumlah]

        if tipe not in templates:
            errors.append(
                f"Tipe '{tipe}' tidak ada di templates.yaml. "
                f"Tipe yang dikenal: {', '.join(sorted(templates))}")
            continue
        try:
            bentang = int(bentang_raw)
        except (TypeError, ValueError):
            errors.append(f"bentang_bersih_mm tidak valid: {bentang_raw!r}")
            continue
        if bentang <= 0:
            errors.append(f"bentang_bersih_mm harus positif: {bentang}")
            continue
        try:
            jumlah = int(jumlah_raw)
        except (TypeError, ValueError):
            errors.append(f"jumlah tidak valid: {jumlah_raw!r}")
            continue
        if jumlah <= 0:
            errors.append(f"jumlah harus positif: {jumlah}")
            continue
        lokasi = str(row[i_lokasi]).strip() if i_lokasi is not None and row[i_lokasi] else ""
        elemen.append(ElemenInput(tipe=tipe, bentang_bersih_mm=bentang,
                                  jumlah=jumlah, lokasi=lokasi))

    if errors:
        raise ConfigError(
            "Input elemen tidak valid.\n\n" + "\n\n".join(
                f"  {e}" for e in errors) +
            "\n\nPerbaiki input/elemen.xlsx lalu jalankan ulang.")
    return elemen
