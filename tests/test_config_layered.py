"""Tests config berlapis (08-SPEC-config-berlapis) — §10.

Wajib: resolve deep merge, dua gambar cover beda → sengkang beda,
migrasi, endpoint drawings, bar mark prefix, asal-nilai.
"""

import sys
import shutil
from pathlib import Path

import pytest
import yaml

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(REPO / "web"))

from config_loader import (list_drawings, load_drawing, load_layered,   # noqa: E402
                           migrate_legacy_layered, resolve_config,
                           load_project_config, load_templates)         # noqa: E402
from models import ConfigError  # noqa: E402
from bbs import generate_bbs, agregasi  # noqa: E402
from models import ElemenInput  # noqa: E402

CONFIG_DIR = REPO / "config"

import app as _appmod


@pytest.fixture(autouse=True)
def _reset_config_dir():
    """Setiap test berlapis pakai tmp CONFIG_DIR; reset ke asli setelahnya."""
    _appmod.CONFIG_DIR = REPO / "config"
    yield
    _appmod.CONFIG_DIR = REPO / "config"


@pytest.fixture(scope="module")
def client():
    from app import app
    app.config["TESTING"] = True
    return app.test_client()


def _buat_proyek_berlapis(tmp_path):
    """Buat proyek berlapis minimal di tmp_path."""
    pdir = tmp_path / "projects" / "LAP1"
    (pdir / "drawings").mkdir(parents=True)
    proj = {
        "proyek": {"nama": "Lapis", "kode": "LAP1"},
        "sumber": {"dokumen": "GS-X", "revisi": "Rev.1", "tanggal": "2026-01-01"},
        "stok": {"panjang_batang_mm": 12000, "kerf_mm": 3, "sisa_min_simpan_mm": 1000},
        "selimut_beton_mm": {"balok": 40, "kolom": 40, "plat": 20},
        "panjang_penyaluran_mm": {10: 400, 13: 520, 19: 760},
        "lap_splice_mm": {},
        "unit_weight_kg_per_m": {10: 0.617, 13: 1.042, 19: 2.226},
        "hook": {"tail_135_mm": {10: 80}, "tail_90_mm": {10: 120},
                 "diameter_bengkok_faktor": 4, "koreksi_bengkokan_aktif": False},
        "sengkang": {"zona_tumpuan_faktor": 0.25, "jarak_sengkang_pertama_mm": 50,
                     "metode_hitung": "kontinyu"},
        "optimizer": {"max_pola": 8, "batasi_pola": False},
    }
    tpl = {"balok": {"B1": {
        "deskripsi": "t", "b_mm": 300, "h_mm": 600,
        "tulangan": [{"posisi": "atas", "dia": 19, "jumlah": 4,
                      "tumpuan_kedua_ujung": True}],
        "sengkang": {"dia": 10, "jarak_tumpuan_mm": 150,
                     "jarak_lapangan_mm": 200, "kaki": 2, "hook_sudut": 135}}}}
    (pdir / "project.yaml").write_text(yaml.safe_dump(proj))
    (pdir / "templates.yaml").write_text(yaml.safe_dump(tpl))
    # gambar GS-01 (kosong) & GS-02 (cover balok 50)
    g1 = {"kode": "GS-01", "nama": "Atas", "revisi": "Rev.1",
          "tanggal": "2026-01-01", "override": {}}
    (pdir / "drawings" / "GS-01.yaml").write_text(yaml.safe_dump(g1))
    g2 = {"kode": "GS-02", "nama": "Basement", "revisi": "Rev.1",
          "tanggal": "2026-01-02",
          "override": {"selimut_beton_mm": {"balok": 50},
                       "panjang_penyaluran_mm": {"19": 820}}}
    (pdir / "drawings" / "GS-02.yaml").write_text(yaml.safe_dump(g2))
    return tmp_path


# ── resolve_config deep merge ──────────────────────────────
def test_resolve_deep_merge(tmp_path):
    base = _buat_proyek_berlapis(tmp_path)
    cfg = load_project_config(base / "projects" / "LAP1" / "project.yaml")
    g2 = load_drawing(base / "projects", "LAP1", "GS-02")
    resolved = resolve_config(cfg, g2["override"])
    # D19 override 820, D10 & D13 tetap dari proyek (deep merge)
    assert resolved.ld[19] == 820
    assert resolved.ld[10] == 400
    assert resolved.ld[13] == 520
    # cover balok 50 (override), kolom/plat tetap 40/20
    assert resolved.cover["balok"] == 50
    assert resolved.cover["kolom"] == 40


# ── dua gambar cover beda → sengkang beda (uji utama) ─────
def test_dua_gambar_sengkang_beda(tmp_path):
    base = _buat_proyek_berlapis(tmp_path)
    cfg1, tpl1, _ = load_layered(base / "projects", "LAP1", "GS-01")
    cfg2, tpl2, _ = load_layered(base / "projects", "LAP1", "GS-02")
    el = ElemenInput(tipe="B1", bentang_bersih_mm=6000, jumlah=1, lokasi="x")
    c1 = generate_bbs(tpl1, [el], cfg1)
    c2 = generate_bbs(tpl2, [el], cfg2)
    sk1 = [c for c in c1 if c.posisi == "sengkang"][0]
    sk2 = [c for c in c2 if c.posisi == "sengkang"][0]
    # GS-01: cover 40 → dalam 220×520 → 1480+160=1640
    # GS-02: cover 50 → dalam 200×500 → 1400+160=1560
    assert sk1.panjang_mm == 1640
    assert sk2.panjang_mm == 1560
    assert sk1.panjang_mm != sk2.panjang_mm
    # tulangan utama D19: GS-01 6000+2×760=7520, GS-02 6000+2×820=7640
    t1 = [c for c in c1 if c.dia == 19][0]
    t2 = [c for c in c2 if c.dia == 19][0]
    assert t1.panjang_mm == 7520 and t2.panjang_mm == 7640


# ── validasi error menyebut gambar ─────────────────────────
def test_validasi_gagal_menyebut_gambar(tmp_path):
    base = _buat_proyek_berlapis(tmp_path)
    # hapus Ld 19 dari proyek → GS-02 yang override 19 tetap 820, tapi B1 butuh
    # 19 → gambar dengan override 19 OK; buat gambar GS-03 tanpa 19 → error
    cfg = load_project_config(base / "projects" / "LAP1" / "project.yaml")
    del cfg.ld[19]
    # tulis ulang proyek tanpa Ld 19
    p = base / "projects" / "LAP1" / "project.yaml"
    d = yaml.safe_load(p.read_text())
    del d["panjang_penyaluran_mm"][19]
    p.write_text(yaml.safe_dump(d))
    # GS-01 (tanpa override 19) → error menyebut GS-01
    with pytest.raises(ConfigError) as exc:
        load_layered(base / "projects", "LAP1", "GS-01")
    assert "GS-01" in str(exc.value)
    # GS-02 override 19 → lolos
    cfg2, _, _ = load_layered(base / "projects", "LAP1", "GS-02")
    assert cfg2.ld[19] == 820


# ── migrasi berlapis ───────────────────────────────────────
def test_migrasi_berlapis(tmp_path):
    cfgdir = tmp_path / "config"
    cfgdir.mkdir()
    (cfgdir / "project.yaml").write_text(
        (CONFIG_DIR / "project.yaml").read_text())
    (cfgdir / "templates.yaml").write_text(
        (CONFIG_DIR / "templates.yaml").read_text())
    ok = migrate_legacy_layered(cfgdir)
    assert ok is True
    pdir = cfgdir / "projects" / "PRJ-001"
    assert (pdir / "project.yaml").exists()
    assert (pdir / "templates.yaml").exists()
    draws = list_drawings(cfgdir / "projects", "PRJ-001")
    assert len(draws) == 1
    assert draws[0]["kode"] == "PRJ-001"
    # migrasi kedua → no-op
    assert migrate_legacy_layered(cfgdir) is False


# ── endpoint drawings ──────────────────────────────────────
def test_endpoint_drawings_crud(client, tmp_path):
    _appmod.CONFIG_DIR = tmp_path
    _buat_proyek_berlapis(tmp_path)
    # GET list
    r = client.get("/api/projects/LAP1/drawings")
    d = r.get_json()
    assert d["ok"] and len(d["drawings"]) == 2
    # GET detail GS-02 → asal nilai
    r = client.get("/api/projects/LAP1/drawings/GS-02")
    d = r.get_json()
    assert d["ok"]
    assert d["asal"]["panjang_penyaluran_mm"]["19"]["asal"] == "gambar"
    assert d["asal"]["panjang_penyaluran_mm"]["10"]["asal"] == "proyek"
    assert d["asal"]["selimut_beton_mm"]["balok"]["asal"] == "gambar"
    # config efektif
    assert d["config_efektif"]["ld"]["19"] == 820
    # POST gambar baru
    r = client.post("/api/projects/LAP1/drawings", json={
        "kode": "GS-03", "nama": "Gedung B", "revisi": "Rev.1",
        "tanggal": "2026-02-01"})
    assert r.status_code == 201
    # PUT override GS-03: cover balok 55 — revisi wajib beda (nilai berubah)
    r = client.put("/api/projects/LAP1/drawings/GS-03", json={
        "override": {"selimut_beton_mm": {"balok": 55}},
        "revisi": "Rev.2"})
    assert r.status_code == 200, r.get_json()
    # revisi sama + nilai berubah → tolak (Rev.2 = yg tersimpan skrg)
    r = client.put("/api/projects/LAP1/drawings/GS-03", json={
        "override": {"selimut_beton_mm": {"balok": 60}},
        "revisi": "Rev.2"})
    assert r.status_code == 400, r.get_json()
    # koreksi checkbox → lolos (revisi tetap Rev.2)
    r = client.put("/api/projects/LAP1/drawings/GS-03", json={
        "override": {"selimut_beton_mm": {"balok": 60}},
        "revisi": "Rev.2", "koreksi_bukan_revisi": True,
        "catatan": "typo"})
    assert r.status_code == 200, r.get_json()
    # yaml unduh
    r = client.get("/api/projects/LAP1/drawings/GS-03/yaml")
    assert r.status_code == 200


# ── hitung pakai proyek+gambar → bar mark prefix ──────────
def test_hitung_bar_mark_prefix(client, tmp_path):
    _appmod.CONFIG_DIR = tmp_path
    _buat_proyek_berlapis(tmp_path)
    r = client.post("/api/hitung", json={
        "proyek": "LAP1", "gambar": "GS-01",
        "elemen": [{"tipe": "B1", "bentang_bersih_mm": 6000, "jumlah": 1}]})
    d = r.get_json()
    assert d["ok"] is True, d
    assert all(b["bar_mark"].startswith("GS-01/") for b in d["bbs"])
    # ganti gambar → hasil beda (cover)
    r2 = client.post("/api/hitung", json={
        "proyek": "LAP1", "gambar": "GS-02",
        "elemen": [{"tipe": "B1", "bentang_bersih_mm": 6000, "jumlah": 1}]})
    d2 = r2.get_json()
    sk1 = [b for b in d["bbs"] if b["posisi"] == "sengkang"][0]
    sk2 = [b for b in d2["bbs"] if b["posisi"] == "sengkang"][0]
    assert sk1["panjang_mm"] != sk2["panjang_mm"]


# ── export pakai proyek+gambar → nama file & header ───────
def test_export_nama_file_dan_header(client, tmp_path):
    import io
    from openpyxl import load_workbook
    _appmod.CONFIG_DIR = tmp_path
    _buat_proyek_berlapis(tmp_path)
    r = client.post("/api/export", json={
        "proyek": "LAP1", "gambar": "GS-02",
        "elemen": [{"tipe": "B1", "bentang_bersih_mm": 6000, "jumlah": 1}]})
    assert r.status_code == 200
    cd = r.headers.get("Content-Disposition", "")
    assert "BBS_LAP1_GS-02" in cd
    wb = load_workbook(io.BytesIO(r.data))
    ws = wb["BBS"]
    texts = [str(ws.cell(rr, 1).value or "") for rr in range(1, 12)]
    assert any("GAMBAR" in t and "GS-02" in t for t in texts)
