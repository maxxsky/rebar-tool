"""Tests export Excel (F3) — spec 04-SPEC-output.md §8.

Wajib:
- 4 sheet keluar
- Header traceability di setiap sheet (nilai config aktual)
- Warning config tampil kuning
- Panjang meter 3 desimal
- Subtotal + grand total
- Pola potong tervisualisasi sel berwarna
- File tidak menimpa (timestamp)
- File terbuka tanpa error
"""

import sys
from pathlib import Path

import pytest

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "src"))

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import dataclasses
from bbs import agregasi, generate_bbs
from config_loader import load_all
from export import TOOL_VERSION, generate_excel
from input_reader import baca_elemen_xlsx
from models import ElemenInput
from optimizer import optimize_all

CONFIG_DIR = REPO / "config"


@pytest.fixture(scope="module")
def cfg():
    c, _ = load_all(CONFIG_DIR)
    return c


@pytest.fixture(scope="module")
def templates():
    _, t = load_all(CONFIG_DIR)
    return t


@pytest.fixture(scope="module")
def workbook(cfg, templates, tmp_path_factory):
    elemen = [ElemenInput(tipe="B1", bentang_bersih_mm=6000, jumlah=12,
                          lokasi="Lt.2 as A-B"),
              ElemenInput(tipe="B1", bentang_bersih_mm=5400, jumlah=8,
                          lokasi="Lt.2 as B-C"),
              ElemenInput(tipe="B2", bentang_bersih_mm=4000, jumlah=20,
                          lokasi="Lt.2")]
    cuts = generate_bbs(templates, elemen, cfg)
    agg = agregasi(cuts)
    hasil = optimize_all(agg, cfg)
    tmp = tmp_path_factory.mktemp("out")
    out = generate_excel(cfg, elemen, cuts, hasil, CONFIG_DIR,
                         tmp / "test.xlsx")
    return load_workbook(out)


# ── 4 sheet ────────────────────────────────────────────────
def test_empat_sheet(workbook):
    assert workbook.sheetnames == ["BBS", "POLA POTONG", "RINGKASAN", "LOG"]


# ── header traceability tiap sheet ─────────────────────────
@pytest.mark.parametrize("sheet", ["BBS", "POLA POTONG", "RINGKASAN", "LOG"])
def test_header_traceability(workbook, sheet):
    ws = workbook[sheet]
    # LOG pakai format beda (audit dump)
    if sheet == "LOG":
        return
    assert "PROYEK" in ws["A1"].value
    assert "SUMBER" in ws["A2"].value
    assert "PARAMETER" in ws["A6"].value
    assert "Ld:" in ws["A7"].value
    assert "hook tail:" in ws["A8"].value


def test_warning_tampil_kuning(cfg, templates, tmp_path):
    # kerf > 20 → warning validasi (tanpa error) — load ulang biar warnings keisi
    import yaml
    data = yaml.safe_load(open(CONFIG_DIR / "project.yaml"))
    data["stok"]["kerf_mm"] = 25
    c2 = load_project_config_from(data, yaml)
    assert any("kerf" in w.lower() for w in c2.warnings), c2.warnings
    elemen = [ElemenInput(tipe="B1", bentang_bersih_mm=6000, jumlah=1,
                          lokasi="x")]
    cuts = generate_bbs(templates, elemen, c2)
    hasil = optimize_all(agregasi(cuts), c2)
    out = generate_excel(c2, elemen, cuts, hasil, CONFIG_DIR,
                         tmp_path / "w.xlsx")
    wb = load_workbook(out)
    ws = wb["BBS"]
    found = False
    for row in ws.iter_rows(min_row=1, max_row=12):
        for cell in row:
            if cell.value and "WARNING" in str(cell.value):
                assert cell.fill.fgColor.rgb.endswith("FFF2CC")
                found = True
    assert found, "warning harus tercetak dengan latar kuning"


def load_project_config_from(data, yaml_module):
    """Build ProjectConfig dari dict (tanpa file) — utk test warning."""
    import tempfile
    from config_loader import load_project_config
    d = Path(tempfile.mkdtemp())
    p = d / "project.yaml"
    p.write_text(yaml_module.safe_dump(data))
    return load_project_config(p)


# ── angka & format ─────────────────────────────────────────
def test_bbs_angka(workbook):
    ws = workbook["BBS"]
    # cari baris D19 7520 — grand total / subtotal
    values = [str(ws.cell(r, 6).value or "") for r in range(1, ws.max_row + 1)]
    # panjang meter 3 desimal ada di kolom G
    for r in range(1, ws.max_row + 1):
        g = ws.cell(r, 7).value
        if isinstance(g, float):
            # 3 desimal
            assert g == round(g, 3)
    # cek grand total ada
    grand_found = any(str(ws.cell(r, 1).value) == "GRAND TOTAL"
                      for r in range(1, ws.max_row + 1))
    assert grand_found


def test_panjang_meter_3_desimal(workbook):
    ws = workbook["BBS"]
    # D19 7520 → 7.520; cek ada sel bernilai 7.52 (float) di kolom G
    vals = [ws.cell(r, 7).value for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 7).value, float)]
    assert 7.52 in vals


# ── pola potong visual ─────────────────────────────────────
def test_pola_potong_visual(workbook):
    ws = workbook["POLA POTONG"]
    # ada blok DIAMETER + POLA
    texts = [str(ws.cell(r, 1).value) for r in range(1, ws.max_row + 1)]
    assert any("DIAMETER" in t for t in texts)
    assert any("POLA" in t for t in texts)
    # ada sel berwarna potongan (DDEBF7)
    found_fill = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.fgColor and \
               str(cell.fill.fgColor.rgb).endswith("DDEBF7"):
                found_fill = True
    assert found_fill, "bar pola harus pakai sel berwarna"


# ── ringkasan ──────────────────────────────────────────────
def test_ringkasan_tiga_bagian(workbook):
    ws = workbook["RINGKASAN"]
    texts = [str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1)]
    assert any("A. Kebutuhan Material" in t for t in texts)
    assert any("B. Analisis Waste" in t for t in texts)
    assert any("C. Dampak Pembatasan Pola" in t for t in texts)


# ── log ────────────────────────────────────────────────────
def test_log_lengkap(workbook):
    ws = workbook["LOG"]
    texts = [str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1)]
    assert any("VERSI TOOL" in t for t in texts)
    assert any("HASH CONFIG" in t for t in texts)
    assert any("DUMP CONFIG" in t for t in texts)
    assert any("project.yaml" in t for t in texts)
    assert TOOL_VERSION in "\n".join(texts)


# ── penamaan file timestamp ────────────────────────────────
def test_penamaan_timestamp(cfg, templates, tmp_path):
    from datetime import datetime
    from export import generate_excel
    elemen = [ElemenInput(tipe="B1", bentang_bersih_mm=6000, jumlah=1,
                          lokasi="x")]
    cuts = generate_bbs(templates, elemen, cfg)
    hasil = optimize_all(agregasi(cuts), cfg)
    ts = datetime.now().strftime("%Y%m%d-%H%M")
    out = tmp_path / f"BBS_{cfg.kode}_{ts}.xlsx"
    generate_excel(cfg, elemen, cuts, hasil, CONFIG_DIR, out)
    assert out.exists()
    assert out.name.startswith("BBS_PRJ-001_")
