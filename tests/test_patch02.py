"""Tests PATCH-02 — panel config editable, write-guard VPS, override luas.

Spec: PATCH-02-config-editable.md §1, §3.
"""

import sys
from pathlib import Path

import pytest
import yaml

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "src"))
sys.path.insert(0, str(REPO / "web"))

from models import ConfigError  # noqa: E402

CONFIG_DIR = REPO / "config"


@pytest.fixture(scope="module")
def client():
    from app import app
    app.config["TESTING"] = True
    return app.test_client()


def _baca_payload(kode="P2A", revisi="Rev.1", ld19=760):
    # payload valid minimal — sama bentuk seperti yg dikirim panel edit
    return {
        "kode": kode,
        "config": {
            "proyek": {"nama": "P2", "kode": kode},
            "sumber": {"dokumen": "GS-P2", "revisi": revisi,
                       "tanggal": "2026-08-01", "catatan": "tes"},
            "stok": {"panjang_batang_mm": 12000, "kerf_mm": 3,
                     "sisa_min_simpan_mm": 1000},
            "selimut_beton_mm": {"balok": 40, "kolom": 40, "plat": 20},
            "panjang_penyaluran_mm": {10: 400, 13: 520, 19: ld19},
            "lap_splice_mm": {},
            "unit_weight_kg_per_m": {10: 0.617, 13: 1.042, 19: 2.226},
            "hook": {"tail_135_mm": {10: 80}, "tail_90_mm": {10: 120},
                     "diameter_bengkok_faktor": 4,
                     "koreksi_bengkokan_aktif": False},
            "sengkang": {"zona_tumpuan_faktor": 0.25,
                         "jarak_sengkang_pertama_mm": 50,
                         "metode_hitung": "kontinyu"},
            "optimizer": {"max_pola": 8, "batasi_pola": False},
        },
        "templates": {
            "balok": {
                "B1": {
                    "deskripsi": "tes", "b_mm": 300, "h_mm": 600,
                    "tulangan": [
                        {"posisi": "atas", "dia": 19, "jumlah": 4,
                         "tumpuan_kedua_ujung": True}],
                    "sengkang": {"dia": 10, "jarak_tumpuan_mm": 150,
                                 "jarak_lapangan_mm": 200, "kaki": 2,
                                 "hook_sudut": 135},
                }
            }
        },
    }


def _setup(kode="P2A", revisi="Rev.1"):
    from app import _simpan_proyek_baru
    _simpan_proyek_baru(_baca_payload(kode, revisi))


def _cleanup(kode="P2A"):
    import shutil
    for d in ("projects", "templates"):
        p = CONFIG_DIR / d / f"{kode}.yaml"
        if p.exists():
            p.unlink()
    for arsip_d in ("_arsip",):
        ad = CONFIG_DIR / arsip_d
        if ad.exists():
            for p in ad.glob("project_*.yaml"):
                p.unlink()
            for p in ad.glob("templates_*.yaml"):
                p.unlink()


# ── PATCH /api/config ──────────────────────────────────────
def test_patch_simpan_config_dan_arsip(client):
    _cleanup(); _setup()
    p = _baca_payload()
    p["config"]["panjang_penyaluran_mm"][19] = 800
    p["config"]["sumber"]["revisi"] = "Rev.2"   # revisi beda → boleh
    r = client.patch("/api/config", json={
        "kode": "P2A", "config": p["config"], "templates": p["templates"]})
    assert r.status_code == 200, r.get_json()
    assert r.get_json()["arsip"], "harus ada file arsip"
    d = yaml.safe_load((CONFIG_DIR / "projects" / "P2A.yaml").read_text())
    ld = {int(k): v for k, v in d["panjang_penyaluran_mm"].items()}
    assert ld[19] == 800
    assert d["_meta"]["diubah_via"] == "web"
    assert d["_meta"]["diubah_dari"].startswith("project_")
    assert list((CONFIG_DIR / "_arsip").glob("project_*.yaml")), "arsip ada"
    _cleanup()


def test_patch_revisi_sama_nilai_berubah_ditolak(client):
    _cleanup(); _setup()
    p = _baca_payload()
    p["config"]["panjang_penyaluran_mm"][19] = 800  # nilai berubah
    # revisi tetap Rev.1 → tolak
    r = client.patch("/api/config", json={
        "kode": "P2A", "config": p["config"], "templates": p["templates"]})
    assert r.status_code == 400
    assert "revisi" in r.get_json()["error"].lower()
    _cleanup()


def test_patch_koreksi_bukan_revisi_dengan_catatan(client):
    _cleanup(); _setup()
    p = _baca_payload()
    p["config"]["panjang_penyaluran_mm"][19] = 800
    r = client.patch("/api/config", json={
        "kode": "P2A", "config": p["config"], "templates": p["templates"],
        "koreksi_bukan_revisi": True, "catatan": "typo di Ld 19"})
    assert r.status_code == 200, r.get_json()
    _cleanup()


def test_patch_koreksi_tanpa_catatan_ditolak(client):
    _cleanup(); _setup()
    p = _baca_payload()
    p["config"]["panjang_penyaluran_mm"][19] = 800
    r = client.patch("/api/config", json={
        "kode": "P2A", "config": p["config"], "templates": p["templates"],
        "koreksi_bukan_revisi": True, "catatan": ""})
    assert r.status_code == 400
    assert "Catatan" in r.get_json()["error"]
    _cleanup()


def test_patch_kerf_saja_tidak_minta_revisi(client):
    """kerf & sisa_min BUKAN nilai teknis (PATCH-02 §1.4) → revisi tidak wajib."""
    _cleanup(); _setup()
    p = _baca_payload()
    p["config"]["stok"]["kerf_mm"] = 5
    # revisi tetap Rev.1 — tapi kerf bukan nilai teknis → harus lolos
    r = client.patch("/api/config", json={
        "kode": "P2A", "config": p["config"], "templates": p["templates"]})
    assert r.status_code == 200, r.get_json()
    _cleanup()


# ── write-guard VPS (opsi 3) ───────────────────────────────
def test_tulis_ditolak_dari_non_localhost(client):
    from app import app as flask_app
    old = flask_app.config.get("TESTING")
    flask_app.config["TESTING"] = False
    try:
        r = client.patch("/api/config", json={"kode": "X"})
        # test client remote_addr default 127.0.0.1 → 400 (kode tidak ada),
        # bukan 403. Untuk simulasi non-localhost pakai environ_base.
        r2 = client.patch(
            "/api/config", json={"kode": "X"},
            environ_base={"REMOTE_ADDR": "1.2.3.4"})
        assert r2.status_code == 403
        assert "localhost" in r2.get_json()["error"]
    finally:
        flask_app.config["TESTING"] = old


# ── override luas + diff ───────────────────────────────────
def test_override_luas_dan_diff(client):
    # PATCH-06 §2: proyek+gambar wajib — pakai PRJ-001/GS-01 (Ld D19=760)
    r = client.post("/api/hitung", json={
        "proyek": "PRJ-001", "gambar": "GS-01",
        "elemen": [{"tipe": "B1", "bentang_bersih_mm": 6000, "jumlah": 1}],
        "override": {"ld": {"10": 400, "13": 520, "19": 1000},
                     "kerf_mm": 5}})
    d = r.get_json()
    assert d["ok"] is True, d
    # panjang D19 = 6000 + 2×1000 = 8000
    panjang19 = [b["panjang_mm"] for b in d["bbs"] if b["dia"] == 19]
    assert 8000 in panjang19
    # diff menyebut Ld 19 & kerf
    txt = "\n".join(d["override_diff"])
    assert "Ld D19" in txt and "760" in txt and "1000" in txt
    assert "kerf_mm" in txt


# ── Excel dengan override → header warning ─────────────────
def test_export_override_warning_header(client, tmp_path):
    import io
    from openpyxl import load_workbook
    r = client.post("/api/export", json={
        "proyek": "PRJ-001", "gambar": "GS-01",
        "elemen": [{"tipe": "B1", "bentang_bersih_mm": 6000, "jumlah": 1}],
        "override": {"kerf_mm": 5}})
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.data))
    ws = wb["BBS"]
    texts = [str(ws.cell(rr, 1).value or "") for rr in range(1, 12)]
    assert any("CONFIG DI-OVERRIDE" in t for t in texts)
    assert any("kerf_mm" in t for t in texts)
    _cleanup()
